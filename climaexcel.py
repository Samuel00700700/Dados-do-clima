import requests as rq
import sys
from openpyxl import load_workbook
from dotenv import load_dotenv
import os

load_dotenv()

api = os.getenv('API_KEY')

url = 'https://api.hgbrasil.com/weather'
    
params = {
    'lat': '-23.5489',
    'lon': '-46.6388',
    'key': api
}

try:
    response = rq.get(url, params=params, timeout=10)
except Exception as e:
    print(f'Erro de conexão: {e}')
    sys.exit(1)

if response.status_code != 200:
    print(f'Erro na requisição: {response.status_code}')
    sys.exit(1)

data = response.json()

if 'results' not in data:
    print('Erro: resposta inesperada da API')
    sys.exit(1)

#Clima de hoje 
cidade = data['results']['city']
descricao = data['results']['description']
umidade = data['results']['humidity']
nuvens = data['results']['cloudiness']
chuva = data['results']['rain']
vento_velo = data['results']['wind_speedy']
vento_dire = data['results']['wind_cardinal']
fuso_hr = data['results']['timezone']

#Previsão para hoje
forecast_hoje = data['results']['forecast'][0]
temperatura_max = forecast_hoje['max']
temperatura_min = forecast_hoje['min']
descricao_forecast = forecast_hoje['description']
dia_sem = forecast_hoje['weekday']
chuva_prob = forecast_hoje.get('rain_probability', 0)
data_dia = forecast_hoje['date']

wb = load_workbook('clima.xlsx')
ws = wb.active
    
ws['A2'] = cidade
ws['B2'] = f'{data_dia} - {dia_sem}'
ws['B4'] = descricao
ws['B5'] = umidade
ws['B6'] = nuvens
ws['B7'] = chuva
ws['B8'] = f'{vento_velo} ({vento_dire})'
ws['D4'] = temperatura_max
ws['D5'] = temperatura_min
ws['D6'] = descricao_forecast
ws['D7'] = chuva_prob

wb.save("clima.xlsx")